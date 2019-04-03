from openpyxl import load_workbook
import os
filePath = os.path.dirname(__file__)
#print(filePath)
kolData = load_workbook(filePath + '\\Data.xlsx')


def loadUrl(platform):   
    kolList = load_workbook(filePath + '\\List.xlsx')
    #获取视频链接列表
    videoList = kolList['%s'%platform]
    #print(videoList.max_row)
    i = 2
    url = []
    while i<=videoList.max_row:
        if videoList['B%d'%i].value == None:
            i+=1
        else:
            url.append(videoList['B%d'%i].value)
            i+=1
    return url

def appendData(data):
    
    videoData = kolData['Data']
    videoData.append((data))
    
def saveData():
    kolData.save(filePath + '\\Data.xlsx')