import datetime,os,time
from openpyxl import load_workbook
import weiboUserInfo,weiboStatusInfo,wanjia,nga,tieba

filePath = os.path.dirname(__file__)

def loadUserIds(userList):
    print('获取%s用户列表：'%userList.title, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    i = 2
    userIds = []
    while i<=userList.max_row:
        if userList['B%d'%i].value == None:
            i+=1
        else:
            userIds.append(userList['B%d'%i].value)
            i+=1
    #print(userIds)
    return userIds

def appendData(data,platform):
    print('已抓取%s数据：'%platform, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    dataWorkbook = load_workbook(filePath + '\\Data.xlsx')
    dataSheet = dataWorkbook['%s'%platform]
    i = 0
    week = datetime.datetime.now().isocalendar()
            #now().isoweekday()
    #print(data)
    while i<len(data):
        timeNow=['',time.strftime('%Y-%m-%d', time.localtime()),time.strftime('%H:%M:%S', time.localtime()),week[1]]
        timeNow.extend(data[i])
        
        dataSheet.append(timeNow)
        i+=1
    #print(data)
    #dataSheet.append((data))
    dataWorkbook.save(filePath + '\\Data.xlsx')
    print('%s数据已保存:'%platform,time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))

def main(m=5):#每m分钟抓取一次
    
    while True:
        while True:
            now = datetime.datetime.now()
            if (now.minute*60+now.hour*3600+now.second)%(m*60) == 0:
                break
            time.sleep(1)
    
        listWorkbook = load_workbook(filePath + '\\List.xlsx')#打开List

        #执行微博
        
        userIds = loadUserIds(listWorkbook['weibo'])#获取微博用户列表
        weiboUserData = weiboUserInfo.weiboUserInfo(userIds)#获取微博用户信息      
        appendData(weiboUserData,'weiboUserInfo')#将数据存入Data

        weiboStatusData = weiboStatusInfo.weiboStatusInfo(userIds)#获取微博内容信息
        appendData(weiboStatusData,'weiboStatus')#将数据存入Data

        #执行玩家电竞社区
        userIds = loadUserIds(listWorkbook['wanjia'])
        wanjiaData = wanjia.wanjia(userIds)
        appendData(wanjiaData,'wanjia')

        #执行NGA
        userIds = loadUserIds(listWorkbook['nga'])
        ngaData = nga.nga(userIds)
        appendData(ngaData,'nga')

        #执行贴吧
        userIds = loadUserIds(listWorkbook['tieba'])
        tiebaData = tieba.tieba(userIds)
        appendData(tiebaData,'tieba')


        time.sleep(50)


main(1)







#TODO

    #计时器： 每到第X小时第X分钟执行

        #打开EXCEL
        #获取id[]
        #执行XXX平台程序  传入id[]  返回data[][]二维列表
        #写入EXCEL

        #执行下一个平台

        #保存EXCEL