import logging
import datetime,os,time
from openpyxl import load_workbook
import touTiao, kuaiShou, youKu, chuShou, xiGua
import weiboUserInfo, weiboStatusInfo, wanjia, nga, tieba, weishi, hupu, v_qq_com, bilibili





def loadUserIds(userList):
    print('获取%s用户列表：' % userList.title, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    i = 2
    userIds = []
    while i <= userList.max_row:
        if userList['B%d' % i].value == None:
            i += 1
        else:
            userIds.append(userList['B%d' % i].value)
            i += 1
    # print(userIds)
    return userIds

def appendData(data, platform):
    print('已抓取%s数据：' % platform, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))
    dataWorkbook = load_workbook(filePath + '\\Data.xlsx')
    dataSheet = dataWorkbook['%s' % platform]
    i = 0
    week = datetime.datetime.now().isocalendar()
    # now().isoweekday()
    # print(data)
    while i < len(data):
        timeNow = ['', time.strftime('%Y-%m-%d', time.localtime()), time.strftime('%H:%M:%S', time.localtime()),
                   week[1]]
        timeNow.extend(data[i])

        dataSheet.append(timeNow)
        i += 1
    # print(data)
    # dataSheet.append((data))
    dataWorkbook.save(filePath + '\\Data.xlsx')
    print('%s数据已保存:' % platform, time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()))


def main(m):
    print('开始')
    while True:
        while True:
            now = datetime.datetime.now()
            if (now.minute * 60 + now.hour * 3600 + now.second) % (m * 60) == 0:
                break
            time.sleep(1)
        listWorkbook = load_workbook(filePath + '\\List.xlsx')  # 打开List
        errmsg = ''
        try:
            errmsg = '快手抓取出错'
            # 执行快手操作
            userIds = loadUserIds(listWorkbook['kuaishou'])  # 获取快手用户列表
            kuaishouUserList = kuaiShou.kuaiShouList(userIds)  # 创建快手用户列表类
            # appendData(kuaishouUserList.getUserInfo(), 'kuaishouUserInfo') # 将数据存入Data，因存在加密字符所以暂不录入
            appendData(kuaishouUserList.getVideoInfo(), 'kuaishouVideo')  # 将数据存入Data

            errmsg = '头条抓取出错'
            # 执行头条操作
            userIds = loadUserIds(listWorkbook['toutiao'])  # 获取头条用户列表
            touTiaoList = touTiao.touTiaoList(userIds)  # 创建头条用户列表类
            appendData(touTiaoList.getUserInfo(), 'toutiaoUserInfo')  # 将数据存入Data
            appendData(touTiaoList.getArtiInfo(), 'toutiaoArticle')  # 将数据存入Data

            errmsg = '优酷抓取出错'
            # 执行优酷
            userIds = loadUserIds(listWorkbook['youku'])  # 获取优酷用户列表
            youKuUserList = youKu.youKuList(userIds)  # 创建优酷用户列表类
            appendData(youKuUserList.getUserInfo(), 'youkuUserInfo')  # 将数据存入Data
            appendData(youKuUserList.getVideoInfo(), 'youkuVideo')  # 将数据存入Data

            errmsg = '西瓜抓取出错'
            # 执行西瓜
            userIds = loadUserIds(listWorkbook['xigua'])
            xiGuaList = xiGua.xiGuaList(userIds)
            appendData(xiGuaList.getUserInfo(), 'xiguaUserInfo')
            appendData(xiGuaList.getVideoInfo(), 'xiguaVideo')

            errmsg = '触手抓取出错'
            # 执行触手
            userIds = loadUserIds(listWorkbook['chushou'])
            chuShouList = chuShou.chuShouList(userIds)
            appendData(chuShouList.getUserInfo(), 'chushouUserInfo')
            appendData(chuShouList.getVideoInfo(), 'chushouVideo')
            appendData(chuShouList.getNewsInfo(), 'chushouNews')

            errmsg = '微博抓取出错'
            # 执行微博
            userIds = loadUserIds(listWorkbook['weibo'])  # 获取微博用户列表
            weiboUserData = weiboUserInfo.weiboUserInfo(userIds)  # 获取微博用户信息
            appendData(weiboUserData, 'weiboUserInfo')  # 将数据存入Data
            weiboStatusData = weiboStatusInfo.weiboStatusInfo(userIds)  # 获取微博内容信息
            appendData(weiboStatusData, 'weiboStatus')  # 将数据存入Data

            errmsg = '玩佳抓取出错'
            # 执行玩家电竞社区
            userIds = loadUserIds(listWorkbook['wanjia'])
            wanjiaData = wanjia.wanjia(userIds)
            appendData(wanjiaData, 'wanjia')

            errmsg = 'NGA抓取出错'
            # 执行NGA
            userIds = loadUserIds(listWorkbook['nga'])
            ngaData = nga.nga(userIds)
            appendData(ngaData, 'nga')

            errmsg = '贴吧抓取出错'
            # 执行贴吧
            userIds = loadUserIds(listWorkbook['tieba'])
            tiebaData = tieba.tieba(userIds)
            appendData(tiebaData, 'tieba')

            errmsg = '微视抓取出错'
            # 执行微视操作
            userIds = loadUserIds(listWorkbook['weishi'])
            weishiData = weishi.weishi(userIds)
            appendData(weishiData[0], 'weishiUserInfo')
            appendData(weishiData[1], 'weishiVideoInfo')

            errmsg = '腾讯视频抓取出错'
            # 腾讯视频
            userIds = loadUserIds(listWorkbook['tengxunshipin'])
            tengxunData = v_qq_com.v_qq(userIds)
            appendData(tengxunData, 'tengxunshipin')

            errmsg = 'B站抓取出错'
            #bilibili
            userIds = loadUserIds(listWorkbook['bilibili'])
            bilibiliData = bilibili.bilibili(userIds)
            appendData(bilibiliData, 'bilibili')

            errmsg = '虎扑抓取出错'
            # hupu
            userIds = loadUserIds(listWorkbook['hupu'])
            hupuData = hupu.hupu(userIds)
            appendData(hupuData, 'hupu')

            print('本次抓取结束')
            time.sleep(50)

        except Exception as e:
            logging.exception(e)
            print(errmsg)
            time.sleep(600)

    # time.sleep(50)


main(30)

# TODO

# 计时器： 每到第X小时第X分钟执行

# 打开EXCEL
# 获取id[]
# 执行XXX平台程序  传入id[]  返回data[][]二维列表
# 写入EXCEL

# 执行下一个平台

# 保存EXCEL