import time, sys,re,datetime
import openpyxl

from danmu import DanMuClient


wb=openpyxl.Workbook()
sheet_chat=wb.create_sheet('Chat')
sheet_noble=wb.create_sheet('Noble')
sheet_chat.append(['NickName','Content','Time'])
sheet_noble.append(['NobleNum','Time'])

def pp(msg):
    print(msg.encode(sys.stdin.encoding, 'ignore').
        decode(sys.stdin.encoding))
rid = input('input RoomId: ')
dmc = DanMuClient('https://www.douyu.com/%s'%rid)
if not dmc.isValid(): print('Url not valid')

@dmc.danmu
def danmu_fn(msg):
    pp('[Chat][%s] %s' % (msg['NickName'], msg['Content']))
    sheet_chat.append([msg['NickName'], msg['Content'],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    

# @dmc.gift
# def gift_fn(msg):
#     pp('[Gift][%s] sent a gift!' % msg['NickName'])

@dmc.bggift
def bggift_fn(msg):
    pp('[Gift][%s] sent %s %s %s!' % (msg['sn'],msg['dn'],msg['gc'],msg['gn']))

@dmc.noble
def noble_fn(msg):
    print('[Noble]: %s'%msg['vn'])
    sheet_noble.append([msg['vn'],datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

# @dmc.bet
# def bet_fn(msg):
#         msg['qril'] = msg['qril'].replace('@A=', ':').replace('/', ',')
#         print(msg)



# @dmc.other
# def other_fn(msg):
#     print(msg)

dmc.start(blockThread = True)
filename=input('输入文件名：')
wb.save('%s.xlsx'%filename)

'''

srres:分享直播间
noble_num_info:贵族信息
spbc：广播火箭


'''
